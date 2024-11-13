from rest_framework import viewsets
from .models import TemperatureHumidityData, SolarRadiationData , RainData
from .serializers import TemperatureHumidityDataSerializer, SolarRadiationDataSerializer , RainDataSerializer
import openpyxl
from django.http import HttpResponse

class TemperatureHumidityDataViewSet(viewsets.ModelViewSet):
    queryset = TemperatureHumidityData.objects.all()
    serializer_class = TemperatureHumidityDataSerializer

class SolarRadiationDataViewSet(viewsets.ModelViewSet):
    queryset = SolarRadiationData.objects.all()
    serializer_class = SolarRadiationDataSerializer

class RainDataViewSet(viewsets.ModelViewSet):
    queryset = RainData.objects.all()
    serializer_class = RainDataSerializer

def export_excel(request): 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Datos Temperatura y Humedad'
    ws.append(['Timestamp', 'Temperatura', 'Humedad'])
    data = TemperatureHumidityData.objects.all()

    ws_solar = wb.create_sheet(title="Datos Radiación Solar")
    ws_solar.append(['Timestamp', 'Radiación Solar (W/m²)'])
    solar_data = SolarRadiationData.objects.all()

    ws_rain = wb.create_sheet(title="Datos Lluvia")
    ws_rain.append(['Timestamp', 'Nivel de Agua', 'Intensidad de Lluvia'])  # Corregido el encabezado
    rain_data = RainData.objects.all()

    for item in data:
        # Remover la zona horaria del timestamp
        timestamp = item.timestamp.replace(tzinfo=None) if item.timestamp else None
        ws.append([timestamp, item.temperature, item.humidity])

    for item in solar_data:
        # Remover la zona horaria del timestamp
        timestamp = item.timestamp.replace(tzinfo=None) if item.timestamp else None
        ws_solar.append([timestamp, item.solar_radiation])

    for item in rain_data:
        # Remover la zona horaria del timestamp
        timestamp = item.timestamp.replace(tzinfo=None) if item.timestamp else None
        ws_rain.append([timestamp, item.water_level, item.rain_intensity])  # Cambiado de ws_solar a ws_rain

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=datos_clima.xlsx'

    wb.save(response)

    return response

